import fitz
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def pdf_to_visual_excel(pdf_path, excel_path):
    doc = fitz.open(pdf_path)
    wb = Workbook()
    
    for page_num in range(len(doc)):
        page = doc.load_page(page_num)
        ws = wb.create_sheet(title=f"Page_{page_num+1}")
        
        # Get all words: (x0, y0, x1, y1, "word", block_no, line_no, word_no)
        words = page.get_text("words")
        
        if not words:
            continue
            
        # Group into rows based on y-coordinate (with a tolerance to handle slight misalignments)
        # Sort by y0 first
        words.sort(key=lambda w: w[1])
        
        rows = []
        current_row = []
        current_y = words[0][1]
        y_tolerance = 5 # point tolerance for same row
        
        for w in words:
            if abs(w[1] - current_y) <= y_tolerance:
                current_row.append(w)
            else:
                rows.append(current_row)
                current_row = [w]
                current_y = w[1]
        if current_row:
            rows.append(current_row)
            
        # Now, figure out global column positions
        # Let's divide page width into discrete logical columns (e.g. 50 point intervals = 1 column)
        col_width = 40
        
        for r_idx, row in enumerate(rows):
            # Sort row by X0
            row.sort(key=lambda w: w[0])
            
            # Combine words that are very close to each other (like first name last name)
            merged_row = []
            if row:
                current_group = row[0]
                x_tolerance = 15 # distance between words to consider them same cell
                for w in row[1:]:
                    if w[0] - current_group[2] <= x_tolerance:
                        # merge
                        current_group = (current_group[0], current_group[1], w[2], max(current_group[3], w[3]), current_group[4] + " " + w[4])
                    else:
                        merged_row.append(current_group)
                        current_group = w
                merged_row.append(current_group)
            
            # Place in excel based on X coordinate
            for item in merged_row:
                col_idx = int(item[0] // col_width) + 1
                
                # if cell already taken, shift right
                while ws.cell(row=r_idx+1, column=col_idx).value is not None:
                    col_idx += 1
                    
                ws.cell(row=r_idx+1, column=col_idx, value=item[4])
                
    if len(wb.sheetnames) > 1:
        del wb['Sheet'] # delete default setup sheet
        
    wb.save(excel_path)
    doc.close()

if __name__ == "__main__":
    pdf_to_visual_excel("test_text.pdf", "test_visual.xlsx")
    print("Test done")
